#!/usr/bin/env python3
"""
Extract raw PDF text structure from the CPA GSL textbook and classify it into
auditable open-book index sheets.

This script intentionally treats PyMuPDF output as a raw extraction layer. It
keeps raw spans and lines, then creates classified index rows with confidence
scores so the result can be reviewed and refined instead of blindly trusted.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from statistics import median
from typing import Any

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PDF_FOLIO_RE = re.compile(r"Pdf_Folio:\s*([0-9ivxlcdmIVXLCDM]+)")
FOOTER_PAGE_RE = re.compile(r"^[0-9ivxlcdmIVXLCDM]{1,8}$")
MODULE_RE = re.compile(r"\bMODULE\s+\d+\b.*", re.IGNORECASE)
FIGURE_RE = re.compile(r"^(FIGURE)\s+([0-9]+(?:\.[0-9]+)+)\b[:.\s-]*(.*)$", re.IGNORECASE)
TABLE_RE = re.compile(r"^(TABLE)\s+([0-9]+(?:\.[0-9]+)+)\b[:.\s-]*(.*)$", re.IGNORECASE)
NUMBERED_HEADING_RE = re.compile(r"^[0-9]+(?:\.[0-9]+){0,3}\s+\S+")
AUTHOR_YEAR_RE = re.compile(
    r"\b([A-Z][A-Za-z'`.-]+(?:\s+(?:and|&)\s+[A-Z][A-Za-z'`.-]+)?)\s*\(((?:19|20)[0-9]{2}[a-z]?)\)"
)
PAREN_CITATION_RE = re.compile(r"\(([A-Z][A-Za-z'`.&,\-\s]{1,90}?,\s*(?:19|20)[0-9]{2}[a-z]?[^)]{0,80})\)")
REFERENCE_LINE_RE = re.compile(r"^([A-Z][A-Za-z'`.-]+(?:,\s*[A-Z][A-Za-z'`.-]+|(?:\s+&\s+|\s+and\s+)[A-Z][A-Za-z'`.-]+){0,8}).{0,160}?\b((?:19|20)[0-9]{2}[a-z]?)\b")
YEAR_RE = re.compile(r"\b((?:19|20)[0-9]{2}[a-z]?)\b")

EXCEL_CELL_LIMIT = 32767


@dataclass
class SpanRow:
    source_span_id: str
    source_line_id: str
    pdf_page: int
    book_page: str
    module: str
    block_no: int
    line_no: int
    span_no: int
    text: str
    font: str
    size: float
    flags: int
    is_bold: bool
    is_italic: bool
    style_type: str
    x0: float
    y0: float
    x1: float
    y1: float


@dataclass
class LineRow:
    source_line_id: str
    pdf_page: int
    book_page: str
    module: str
    block_no: int
    line_no: int
    text: str
    max_size: float
    median_size: float
    font_summary: str
    is_bold: bool
    is_italic: bool
    style_summary: str
    x0: float
    y0: float
    x1: float
    y1: float


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def excel_text(value: Any) -> Any:
    if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
        return value[: EXCEL_CELL_LIMIT - 20] + " [TRUNCATED]"
    return value


def rounded(value: float) -> float:
    return round(float(value), 2)


def is_bold_font(font: str, flags: int) -> bool:
    f = font.lower()
    return bool(flags & 16) or any(token in f for token in ("bold", "-bd", "bd", "heavy", "-hv", "black", "-md", "medium"))


def is_italic_font(font: str, flags: int) -> bool:
    f = font.lower()
    return bool(flags & 2) or any(token in f for token in ("italic", "-it", "ltit", "oblique"))


def style_type(is_bold: bool, is_italic: bool) -> str:
    if is_bold and is_italic:
        return "Bold & Italic"
    if is_bold:
        return "Bold"
    if is_italic:
        return "Italic"
    return ""


def detect_book_page(lines: list[dict[str, Any]]) -> str:
    for line in lines:
        match = PDF_FOLIO_RE.search(line["text"])
        if match:
            return match.group(1)

    footer_candidates = []
    for line in lines:
        text = line["text"].strip()
        if line["y0"] >= 730 and FOOTER_PAGE_RE.match(text):
            footer_candidates.append((line["x0"], text))
    if footer_candidates:
        footer_candidates.sort(key=lambda item: item[0])
        return footer_candidates[0][1]
    return ""


def detect_page_module(lines: list[dict[str, Any]], previous_module: str) -> str:
    footer_modules = [line["text"] for line in lines if line["y0"] >= 745 and MODULE_RE.search(line["text"])]
    if footer_modules:
        return clean_text(footer_modules[-1])
    top_modules = [line["text"] for line in lines if line["y0"] <= 140 and line["text"].strip().startswith("MODULE ")]
    if top_modules:
        return clean_text(top_modules[0])
    return previous_module


def is_footer_or_hidden(line: dict[str, Any]) -> bool:
    text = line["text"]
    if not text:
        return True
    if PDF_FOLIO_RE.search(text):
        return True
    if line["max_size"] < 1:
        return True
    if line["y0"] >= 745:
        return True
    return False


def first_pass_extract(pdf_path: Path, max_pages: int | None = None) -> tuple[list[LineRow], list[SpanRow], dict[str, Any]]:
    doc = fitz.open(pdf_path)
    page_count = len(doc) if max_pages is None else min(max_pages, len(doc))
    all_lines: list[LineRow] = []
    all_spans: list[SpanRow] = []
    size_by_chars: Counter[float] = Counter()
    line_size_counter: Counter[float] = Counter()
    current_module = ""

    for page_index in range(page_count):
        page = doc[page_index]
        pdf_page = page_index + 1
        raw_page_lines: list[dict[str, Any]] = []
        raw_page_spans: list[dict[str, Any]] = []
        line_counter = 1

        for block_no, block in enumerate(page.get_text("dict").get("blocks", []), start=1):
            if "lines" not in block:
                continue
            for local_line_no, line in enumerate(block.get("lines", []), start=1):
                spans = []
                for span_no, span in enumerate(line.get("spans", []), start=1):
                    text = clean_text(span.get("text", ""))
                    if not text:
                        continue
                    bbox = span.get("bbox", [0, 0, 0, 0])
                    size = rounded(span.get("size", 0))
                    flags = int(span.get("flags", 0))
                    font = span.get("font", "")
                    bold = is_bold_font(font, flags)
                    italic = is_italic_font(font, flags)
                    spans.append(
                        {
                            "span_no": span_no,
                            "text": text,
                            "font": font,
                            "size": size,
                            "flags": flags,
                            "is_bold": bold,
                            "is_italic": italic,
                            "style_type": style_type(bold, italic),
                            "bbox": [rounded(v) for v in bbox],
                        }
                    )
                    if size >= 1:
                        size_by_chars[round(size, 1)] += len(text)
                if not spans:
                    continue

                line_text = clean_text("".join(span["text"] for span in spans))
                if not line_text:
                    continue
                line_bbox = line.get("bbox", [0, 0, 0, 0])
                sizes = [span["size"] for span in spans]
                fonts = Counter(span["font"] for span in spans)
                line_id = f"P{pdf_page:03d}-L{line_counter:03d}"
                raw_line = {
                    "source_line_id": line_id,
                    "pdf_page": pdf_page,
                    "book_page": "",
                    "module": "",
                    "block_no": block_no,
                    "line_no": line_counter,
                    "text": line_text,
                    "max_size": rounded(max(sizes)),
                    "median_size": rounded(median(sizes)),
                    "font_summary": "; ".join(f"{font} ({count})" for font, count in fonts.most_common()),
                    "is_bold": any(span["is_bold"] for span in spans),
                    "is_italic": any(span["is_italic"] for span in spans),
                    "style_summary": ", ".join(sorted({span["style_type"] for span in spans if span["style_type"]})),
                    "x0": rounded(line_bbox[0]),
                    "y0": rounded(line_bbox[1]),
                    "x1": rounded(line_bbox[2]),
                    "y1": rounded(line_bbox[3]),
                }
                raw_page_lines.append(raw_line)
                for span in spans:
                    raw_page_spans.append(
                        {
                            **span,
                            "source_span_id": f"{line_id}-S{span['span_no']:02d}",
                            "source_line_id": line_id,
                            "pdf_page": pdf_page,
                            "book_page": "",
                            "module": "",
                            "block_no": block_no,
                            "line_no": line_counter,
                        }
                    )
                if raw_line["max_size"] >= 1:
                    line_size_counter[round(raw_line["max_size"], 1)] += 1
                line_counter += 1

        book_page = detect_book_page(raw_page_lines)
        current_module = detect_page_module(raw_page_lines, current_module)

        for line in raw_page_lines:
            line["book_page"] = book_page
            line["module"] = current_module
            all_lines.append(LineRow(**line))
        for span in raw_page_spans:
            bbox = span.pop("bbox")
            span["book_page"] = book_page
            span["module"] = current_module
            all_spans.append(
                SpanRow(
                    source_span_id=span["source_span_id"],
                    source_line_id=span["source_line_id"],
                    pdf_page=span["pdf_page"],
                    book_page=span["book_page"],
                    module=span["module"],
                    block_no=span["block_no"],
                    line_no=span["line_no"],
                    span_no=span["span_no"],
                    text=span["text"],
                    font=span["font"],
                    size=span["size"],
                    flags=span["flags"],
                    is_bold=span["is_bold"],
                    is_italic=span["is_italic"],
                    style_type=span["style_type"],
                    x0=bbox[0],
                    y0=bbox[1],
                    x1=bbox[2],
                    y1=bbox[3],
                )
            )

    body_candidates = Counter({size: chars for size, chars in size_by_chars.items() if 8.5 <= size <= 11.0})
    body_size = body_candidates.most_common(1)[0][0] if body_candidates else 10.0
    profile = {
        "pdf_path": str(pdf_path),
        "pdf_pages_processed": page_count,
        "body_size_estimate": body_size,
        "span_size_top_20": "; ".join(f"{size}: {count}" for size, count in size_by_chars.most_common(20)),
        "line_size_top_20": "; ".join(f"{size}: {count}" for size, count in line_size_counter.most_common(20)),
    }
    return all_lines, all_spans, profile


def table_or_figure_caption(lines: list[LineRow], index: int) -> str:
    current = lines[index]
    caption = current.text
    for next_line in lines[index + 1 : index + 4]:
        if next_line.pdf_page != current.pdf_page:
            break
        same_row = abs(next_line.y0 - current.y0) <= 3 and next_line.x0 > current.x0
        near_next = 0 < next_line.y0 - current.y0 <= 24 and next_line.x0 >= current.x0
        if (same_row or near_next) and not FIGURE_RE.match(next_line.text) and not TABLE_RE.match(next_line.text):
            if len(next_line.text) <= 180:
                caption = clean_text(f"{caption} {next_line.text}")
                break
    return caption


def classify_heading(line: LineRow, body_size: float) -> tuple[str, str, float] | None:
    text = line.text
    if is_footer_or_hidden(asdict(line)):
        return None
    if FIGURE_RE.match(text) or TABLE_RE.match(text):
        return None
    if len(text) > 220:
        return None

    if MODULE_RE.search(text) and line.y0 < 160:
        return ("Heading", "Module Heading", 0.96)
    if line.max_size >= body_size + 6:
        return ("Heading", "Main/Module Heading", 0.95)
    if line.max_size >= body_size + 3:
        return ("Heading", "Main Heading", 0.9)
    if NUMBERED_HEADING_RE.match(text) and (line.max_size >= body_size + 1.5 or line.is_bold):
        return ("Heading", "Section Heading", 0.9)
    if line.max_size >= body_size + 1.0 and line.is_bold:
        return ("Heading", "Section Heading", 0.85)
    if line.is_bold and len(text) <= 90 and not text.endswith(".") and not text.startswith("•"):
        return ("Heading", "Subheading Candidate", 0.68)
    return None


def extract_citations(text: str) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for match in AUTHOR_YEAR_RE.finditer(text):
        output.append(
            {
                "citation_text": match.group(0),
                "author_string": match.group(1),
                "year": match.group(2),
                "citation_type": "Author-Year",
            }
        )
    for match in PAREN_CITATION_RE.finditer(text):
        value = match.group(1)
        year = YEAR_RE.search(value)
        author = re.split(r",\s*(?:19|20)[0-9]{2}", value, maxsplit=1)[0].strip()
        output.append(
            {
                "citation_text": f"({value})",
                "author_string": author,
                "year": year.group(1) if year else "",
                "citation_type": "Parenthetical",
            }
        )
    return output


def split_author_names(author_string: str) -> list[str]:
    cleaned = re.sub(r"\bet\s+al\.?", "", author_string, flags=re.IGNORECASE)
    parts = re.split(r"\s+(?:and|&)\s+|,\s*", cleaned)
    return [part.strip(" .;:") for part in parts if len(part.strip(" .;:")) >= 2]


def classify(lines: list[LineRow], spans: list[SpanRow], body_size: float) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    master: list[dict[str, Any]] = []
    current_heading = ""
    heading_by_line_id: dict[str, str] = {}
    rows_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)

    def add_row(category: str, element_type: str, content: str, line: LineRow | None, span: SpanRow | None, confidence: float, extra: dict[str, Any] | None = None) -> None:
        source = line or span
        assert source is not None
        row = {
            "Row ID": f"M{len(master) + 1:06d}",
            "Category": category,
            "Element Type": element_type,
            "Content/Text": clean_text(content),
            "Module": getattr(source, "module", ""),
            "Current Heading": current_heading,
            "Book Page": getattr(source, "book_page", ""),
            "PDF Page": getattr(source, "pdf_page", ""),
            "Source Line ID": getattr(source, "source_line_id", ""),
            "Source Span ID": getattr(source, "source_span_id", ""),
            "Confidence": confidence,
            "Font/Size": f"{getattr(source, 'font', getattr(source, 'font_summary', ''))} / {getattr(source, 'size', getattr(source, 'max_size', ''))}",
            "Notes": "",
        }
        if extra:
            row.update(extra)
        master.append(row)
        rows_by_type[category].append(row)

    lines_by_page: dict[int, list[LineRow]] = defaultdict(list)
    for line in lines:
        lines_by_page[line.pdf_page].append(line)

    for pdf_page in sorted(lines_by_page):
        page_lines = sorted(lines_by_page[pdf_page], key=lambda item: (item.y0, item.x0))
        for index, line in enumerate(page_lines):
            line_dict = asdict(line)
            if is_footer_or_hidden(line_dict):
                continue
            heading_by_line_id[line.source_line_id] = current_heading

            figure_match = FIGURE_RE.match(line.text)
            table_match = TABLE_RE.match(line.text)
            is_actual_figure_label = bool(figure_match and line.text.startswith("FIGURE "))
            is_actual_table_label = bool(table_match and line.text.startswith("TABLE "))
            if is_actual_figure_label:
                caption = table_or_figure_caption(page_lines, index)
                add_row("Figure", "Figure Heading", caption, line, None, 0.94, {"Label": figure_match.group(0).split()[0] + " " + figure_match.group(2)})
                continue
            if is_actual_table_label:
                caption = table_or_figure_caption(page_lines, index)
                add_row("Table", "Table Heading", caption, line, None, 0.94, {"Label": table_match.group(0).split()[0] + " " + table_match.group(2)})
                continue

            heading = classify_heading(line, body_size)
            if heading:
                category, element_type, confidence = heading
                add_row(category, element_type, line.text, line, None, confidence)
                current_heading = line.text

            for citation in extract_citations(line.text):
                add_row(
                    "Citation",
                    citation["citation_type"],
                    citation["citation_text"],
                    line,
                    None,
                    0.78,
                    {"Author String": citation["author_string"], "Year": citation["year"]},
                )

            is_reference_section = current_heading.strip().upper() == "REFERENCES"
            reference_match = REFERENCE_LINE_RE.match(line.text)
            if is_reference_section and reference_match and line.max_size <= body_size + 0.5:
                add_row(
                    "Reference Author",
                    "Reference-like Author String",
                    reference_match.group(1),
                    line,
                    None,
                    0.62,
                    {"Author String": reference_match.group(1), "Year": reference_match.group(2)},
                )

    for span in spans:
        if span.style_type and span.size >= 1 and span.y0 < 745 and not PDF_FOLIO_RE.search(span.text):
            add_row(
                "Typography",
                span.style_type,
                span.text,
                None,
                span,
                0.82,
                {"Current Heading": heading_by_line_id.get(span.source_line_id, "")},
            )

    author_summary: dict[str, dict[str, Any]] = {}
    for row in master:
        if row["Category"] not in {"Citation", "Reference Author"}:
            continue
        author_string = row.get("Author String", "")
        for author in split_author_names(author_string):
            summary = author_summary.setdefault(
                author,
                {
                    "Author Name": author,
                    "Occurrences": 0,
                    "Years": set(),
                    "Book Pages": set(),
                    "PDF Pages": set(),
                    "Source Types": set(),
                },
            )
            summary["Occurrences"] += 1
            if row.get("Year"):
                summary["Years"].add(str(row["Year"]))
            if row.get("Book Page"):
                summary["Book Pages"].add(str(row["Book Page"]))
            if row.get("PDF Page"):
                summary["PDF Pages"].add(str(row["PDF Page"]))
            summary["Source Types"].add(str(row["Category"]))

    authors_rows = []
    for value in author_summary.values():
        authors_rows.append(
            {
                "Author Name": value["Author Name"],
                "Occurrences": value["Occurrences"],
                "Years": ", ".join(sorted(value["Years"])),
                "Book Pages": ", ".join(sorted(value["Book Pages"], key=lambda x: (not x.isdigit(), x))),
                "PDF Pages": ", ".join(sorted(value["PDF Pages"], key=lambda x: int(x) if x.isdigit() else 999999)),
                "Source Types": ", ".join(sorted(value["Source Types"])),
            }
        )
    rows_by_type["Author Index"] = sorted(authors_rows, key=lambda row: (-row["Occurrences"], row["Author Name"].lower()))
    return master, rows_by_type


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="13384D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1E8")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        sample = [str(cell.value or "") for cell in column_cells[:250]]
        width = min(max([len(header)] + [len(value) for value in sample]) + 2, 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(columns)
    for row in rows:
        ws.append([excel_text(row.get(column, "")) for column in columns])
    style_sheet(ws)


def write_outputs(
    output_xlsx: Path,
    output_csv: Path,
    lines: list[LineRow],
    spans: list[SpanRow],
    profile: dict[str, Any],
    master: list[dict[str, Any]],
    rows_by_type: dict[str, list[dict[str, Any]]],
) -> None:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme.append(["Field", "Value"])
    for key, value in profile.items():
        readme.append([key, value])
    readme.append(["Classified rows", len(master)])
    readme.append(["Raw lines", len(lines)])
    readme.append(["Raw spans", len(spans)])
    readme.append(["Caveat", "This workbook is generated from PDF font/layout extraction. Review critical entries against the textbook."])
    style_sheet(readme)

    master_columns = [
        "Row ID",
        "Category",
        "Element Type",
        "Content/Text",
        "Module",
        "Current Heading",
        "Book Page",
        "PDF Page",
        "Source Line ID",
        "Source Span ID",
        "Confidence",
        "Font/Size",
        "Author String",
        "Year",
        "Label",
        "Notes",
    ]
    write_sheet(wb, "Classified_Master", master, master_columns)

    line_columns = list(asdict(lines[0]).keys()) if lines else []
    span_columns = list(asdict(spans[0]).keys()) if spans else []
    write_sheet(wb, "Raw_Lines", [asdict(row) for row in lines], line_columns)
    write_sheet(wb, "Raw_Spans", [asdict(row) for row in spans], span_columns)

    headings = [row for row in master if row["Category"] == "Heading"]
    section_headings = [row for row in headings if "Section" in row["Element Type"] or "Subheading" in row["Element Type"]]
    write_sheet(wb, "Headings_All", headings, master_columns)
    write_sheet(wb, "Section_Headings", section_headings, master_columns)
    write_sheet(wb, "Table_Headings", rows_by_type.get("Table", []), master_columns)
    write_sheet(wb, "Figure_Headings", rows_by_type.get("Figure", []), master_columns)
    write_sheet(wb, "Styled_Text", rows_by_type.get("Typography", []), master_columns)
    write_sheet(wb, "Citations", rows_by_type.get("Citation", []), master_columns)
    write_sheet(wb, "Reference_Authors", rows_by_type.get("Reference Author", []), master_columns)
    write_sheet(
        wb,
        "Authors_Index",
        rows_by_type.get("Author Index", []),
        ["Author Name", "Occurrences", "Years", "Book Pages", "PDF Pages", "Source Types"],
    )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=master_columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(master)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract and classify GSL PDF index data.")
    parser.add_argument("--pdf", default="/Users/nuppalap/Desktop/quiz/cpa - gsl.pdf", help="Path to the source PDF.")
    parser.add_argument("--output-xlsx", default="/Users/nuppalap/Desktop/quiz/gsl_pdf_extraction_master.xlsx", help="Output workbook path.")
    parser.add_argument("--output-csv", default="/Users/nuppalap/Desktop/quiz/gsl_pdf_extraction_master.csv", help="Output master CSV path.")
    parser.add_argument("--max-pages", type=int, default=None, help="Optional page limit for testing.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    pdf_path = Path(args.pdf)
    output_xlsx = Path(args.output_xlsx)
    output_csv = Path(args.output_csv)

    lines, spans, profile = first_pass_extract(pdf_path, args.max_pages)
    master, rows_by_type = classify(lines, spans, float(profile["body_size_estimate"]))
    write_outputs(output_xlsx, output_csv, lines, spans, profile, master, rows_by_type)

    print(f"PDF pages processed: {profile['pdf_pages_processed']}")
    print(f"Estimated body size: {profile['body_size_estimate']}")
    print(f"Raw lines: {len(lines)}")
    print(f"Raw spans: {len(spans)}")
    print(f"Classified rows: {len(master)}")
    print(f"Headings: {len(rows_by_type.get('Heading', []))}")
    print(f"Tables: {len(rows_by_type.get('Table', []))}")
    print(f"Figures: {len(rows_by_type.get('Figure', []))}")
    print(f"Styled text: {len(rows_by_type.get('Typography', []))}")
    print(f"Citations: {len(rows_by_type.get('Citation', []))}")
    print(f"Reference authors: {len(rows_by_type.get('Reference Author', []))}")
    print(f"Author index rows: {len(rows_by_type.get('Author Index', []))}")
    print(f"Wrote workbook: {output_xlsx}")
    print(f"Wrote CSV: {output_csv}")


if __name__ == "__main__":
    main()
